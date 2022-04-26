import openpyxl
import os
import random
import subprocess

#Gets Windows language
output = subprocess.check_output(['powershell.exe' ,"[CultureInfo]::InstalleduICulture.Name"], universal_newlines=True)
key = output.split()

CHARS = 'abcdefghijklnop1234567qrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890abcde'
password = ""
wb = openpyxl.load_workbook('exel.xlsx')
wb1 = openpyxl.Workbook()
if (os.path.exists("Passwaord.xlsx")):
    os.remove("Password.xlsx")
wb1.create_sheet(title = 'Password', index = 0)

sheet_users = wb['Users']
sheet_group = wb['Group']
sheet_password = wb1['Password']

print("Enter number of users:")
rows_2 = int(input())
cols_2 = sheet_users.max_column

print("Enter number of groups:")
rows_1 = int(input())
cols_1 = sheet_group.max_column

print("Enter length of password")
chars_password = int(input())

print("Add users to \"Remote Desktop Users\" group? /Y or /N")
choice = input()

#Creating groups
for i in range(2, rows_1 + 2):
    for j in range(1, cols_1 + 1):
        if(j == 1):
            group_name = sheet_group.cell(row = i, column = j).value
        elif(j == 2):
            description_group = sheet_group.cell(row = i, column = j).value
    os.system(f"net localgroup \"{group_name}\" /add")
    os.system(f"net localgroup \"{group_name}\" /comment:\"{description_group}\"")


#Creating users and addind groups
for n in range(2,rows_2 + 2):
    #Generating password
    for k in range(1, chars_password):
            password += random.choice(CHARS)

    for m in range(1, cols_2 + 1):
        if(m == 1):
            user_name = sheet_users.cell(row = n,column = m).value
            sheet_password.cell(row = n - 1, column = m).value = f"{user_name}"
            os.system(f"net user \"{user_name}\" {password} /ADD /yes")
            os.system(f"wmic UserAccount where Name=\"{user_name}\" set PasswordExpires=False")
        elif(m == 2):
            full_name = sheet_users.cell(row = n,column = m).value
            #Saving password in file
            sheet_password.cell(row = n - 1, column = m).value = f"{password}"
            os.system(f"net user \"{user_name}\" /fullname:\"{full_name}\"")
        elif(m == 3):
            description = sheet_users.cell(row = n,column = m).value
            os.system(f"net user \"{user_name}\" /comment:\"{description}\"")
        elif(m == 4):
            group_list = sheet_users.cell(row = n,column = m).value
            group_add = group_list.split(",")
            for i in range(len(group_add)):
                os.system(f"net localgroup \"{group_add[i]}\" \"{user_name}\" /add")
            if(choice == "/Y" or choice == "/y"):
                if (key[0] == "uk-UA" or key[0] == "en-US"):
                    remote_group = "Remote Desktop Users"
                    os.system(f"net localgroup \"{remote_group}\" \"{user_name}\" /add")
                else:
                    remote_group = "Пользователи удаленного рабочего стола"
                    os.system(f"net localgroup \"{remote_group}\" \"{user_name}\" /add")
            else:
                pass
    password = ""   


  
if (os.path.exists("Passwaord.xlsx")):
    os.remove("Password.xlsx")
    
wb1.save('Password.xlsx')

exit = input()

